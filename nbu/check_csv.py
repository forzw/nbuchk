#!/usr/bin/env python
# _*_ coding:utf-8 _*_
#
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# write by : kevin chen
# version : 1.0
# update : 2024-12-01 create scripts
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
#

import csv
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Side, Border
import subprocess as sb
from nbu.check_env import *

def check_csv_excel(args):

    logger.info('start convert csv to excel')

    wb = openpyxl.Workbook()
    #ws = wb.create_sheet(args.master)
    ws = wb.active
    title = ['备份域', '任务ID', '任务类型', '备份ID', '策略名字', '策略类型','客户端名字', '计划名字', '备份数据(KB)', '开始时间', '结束时间', '备份状态', '状态描述']
    ws.append(title)
    ws.auto_filter.ref = 'A1:M1'
    
    with open(args.jobs_csv) as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            print(row)
            ws.append(row)

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 10
    ws.column_dimensions['M'].width = 10

    # 细框线
    side = Side('thin')
    # 边框样式 底边 右边
    border = Border(bottom=side,  right=side)

    # 雾玫瑰色
    header_fill = PatternFill('solid', fgColor='FFE4E1')
    # 深海绿1色
    bottom_fill = PatternFill('solid', fgColor='C1FFC1')

    # 水平和垂直方向居中
    align = Alignment(horizontal='center', vertical='center')

    # 表头设置
    for cell in ws[1]:  # 需要得到单元格对象cell
        # 单元格添加右下细框线
        cell.border = border
        # 单元格添加雾玫瑰色填充
        cell.fill = header_fill
        # 单元格添加对齐
        cell.alignment = align

    # 从第二行开始设置
    for row in ws.iter_rows(min_row=2):
        for cell in row:    # 需要得到单元格对象cell
            # 单元格添加右下细框线
            cell.border = border
            # 单元格添加深海绿1色填充
            cell.fill = bottom_fill
            # 单元格添加对齐
        cell.alignment = align

    wb.save(args.report_xlsx)

