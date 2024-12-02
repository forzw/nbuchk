#!/usr/bin/env python
# _*_ coding:utf-8 _*_
#
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# write by : kevin chen
# version : 1.0
# update : 2024-12-01 create scripts
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
#

import csv
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Side, Border,Font
import subprocess as sb
# from nbu.check_env import *
import os,sys
import datetime
def is_today(date_str):
    try:
        date_obj = datetime.datetime.strptime(date_str, "%Y%m%d").date()
        today = datetime.date.today()
        return date_obj == today
    except ValueError:
        return False
    
def get_csv_data(csv_path):
    ret_json = {}
    for filename in os.listdir(csv_path):
        if filename.endswith(".csv"):
            tmp_name = filename.replace(".csv","")
            name_list = tmp_name.split("_")
            if len(name_list) < 3:
                continue
            #find current day
            date_str = name_list[-1]
            if not is_today(date_str):
                continue
            master_name = name_list[0]
            type_name = name_list[1]
            if master_name not in ret_json:
                ret_json[master_name] = {}
                ret_json[master_name][type_name] = []
            else:
                ret_json[master_name][type_name] = []
                
            filepath = os.path.join(csv_path, filename)
            with open(filepath, 'r', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                # 遍历每一行数据
                for row in reader:
                    # 处理每行数据，例如打印或进行其他操作
                    ret_json[master_name][type_name].append(row)
    #print('get all data done')
    return ret_json

def set_column_titles(ws, row_index,column_titles):
    """
    设置工作表列标题的格式。

    参数：
        ws: 工作表对象。
        column_titles: 列标题列表。
    """
    for col_num, title in enumerate(column_titles, 1):
        cell = ws.cell(row=row_index, column=col_num)
        cell.value = title

        # 设置字体
        font = Font(name="宋体", size=11, bold=True, color="00C0C0C0")  # RGB(191, 191, 191)
        cell.font = font

        # 设置填充颜色
        fill = PatternFill(fill_type="solid", fgColor="00C0C0C0")  # RGB(191, 191, 191)
        cell.fill = fill

        # 设置对齐方式
        alignment = Alignment(horizontal="center", vertical="center")
        cell.alignment = alignment


def generate_job_report(wb,src_data):
    for master_name,sub_data in src_data.items():
        new_sheet = wb.create_sheet(master_name)
        if "jobs" in sub_data:
            new_sheet.append(['备份域','任务ID', '任务类型', '备份ID', '策略名字', '策略类型', '客户端名字','计划名字', '备份数据', '开始时间','结束时间', '备份状态','状态解释'])
            for job in sub_data["jobs"]:
                new_sheet.append(job)
            adjust_column_widths(new_sheet)
            render_jobs(new_sheet)

def generate_sub_report(sheet,src_data,sub_key):
    for _,sub_data in src_data.items():
        if sub_key in sub_data:
            for sub_value in sub_data[sub_key]:
                sheet.append(sub_value)

    sheet.append([])
    sheet.append([])

def adjust_column_widths(sheet):
    """
    自动调整工作表中所有列的宽度以适应内容。

    Args:
        sheet: 要调整列宽的工作表对象。
    """
    for col in sheet.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            except:
                pass
        sheet.column_dimensions[col[0].column_letter].width = max_length + 2

def add_border(cell):
    #border
    border = Border(left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    cell.border = border

def add_fill_color(cell):
    fill = PatternFill(fill_type="solid", fgColor="FFFF00")  # RGB(255,255, 0)
    cell.fill = fill 
                   
def render_sheet(sheet):
    space_rows = 0
    type_index = 0
    is_title = True
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        if is_title:
            for cell in row:
                if not cell.value:
                    continue
                # 设置字体
                font = Font(name="宋体", size=11, bold=True)  # RGB(191, 191, 191)
                cell.font = font

                # 设置填充颜色
                fill = PatternFill(fill_type="solid", fgColor="00C0C0C0")  # RGB(191, 191, 191)
                cell.fill = fill
                #border
                add_border(cell)

            space_rows = 0
            is_title = False
        else:
            if not row[0].value:
                space_rows += 1
                if space_rows == 2:
                    is_title = True
                    type_index += 1
                continue
            for cell in row:
                #border
                if cell.value:
                    add_border(cell)  
            #sum:[备份失败,正在运行]大于0,标记,成功率<90,标记
            #tape:可用率<80,标记
            #disk:使用率>80,标记   
            if type_index == 0:#sum
               try:
                cell = row[8] #backup fail
                if int(cell.value) > 0:
                    add_fill_color(cell)
                cell = row[9] #backup run
                if int(cell.value) > 0:
                    add_fill_color(cell)
                cell = row[10] #success rate
                if float(cell.value) < 90:
                    add_fill_color(cell)
               except (ValueError) as e:
                   continue
            if type_index == 1:#tape
               try:
                cell = row[6] #avaliable rate
                if float(cell.value) < 80:
                    add_fill_color(cell)
               except (ValueError) as e:
                   continue            
            if type_index == 2:#disk
               try:
                cell = row[5] #avaliable rate
                if float(cell.value) > 80:
                    add_fill_color(cell)
               except (ValueError) as e:
                   continue 

def render_jobs(sheet):
    is_title = True
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        if is_title:
            for cell in row:
                if not cell.value:
                    continue
                # 设置字体
                font = Font(name="宋体", size=11, bold=True)  # RGB(191, 191, 191)
                cell.font = font

                # 设置填充颜色
                fill = PatternFill(fill_type="solid", fgColor="00C0C0C0")  # RGB(191, 191, 191)
                cell.fill = fill
                #border
                add_border(cell)
            is_title = False
        else:
            for cell in row:
                #border
                if cell.value:
                    add_border(cell)  
            try:
                cell = row[-2] #job status
                if cell.value and int(cell.value) != 0:
                    add_fill_color(cell)
                cell = row[-1] #description
                if cell.value and (not cell.value.lower().startswith("done")):
                    add_fill_color(cell)
            except (ValueError) as e:
                continue


def generate_report(src_data,filename):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "index"
    #sum
    sheet.append(['备份域', '开始时间', '结束时间', 'Media数量', 'Client数量', '策略数量','备份总数', '备份成功', '备份失败', '正在运行', '成功率'])
    generate_sub_report(sheet,src_data,"sum")

    sheet.append(['备份域', '磁带池', '总数量', '已满', '已冻结', '可使用','可用率'])
    generate_sub_report(sheet,src_data,"tape")

    sheet.append(['备份域', '磁盘备份池', '总容量(TB)', '已使用(TB)', '未使用(TB)', '使用率'])
    generate_sub_report(sheet,src_data,"disk")

    sheet.append(['备份域', '24小时无备份任务策略列表'])
    generate_sub_report(sheet,src_data,"policy")
    render_sheet(sheet)

    adjust_column_widths(sheet)
    generate_job_report(wb,src_data)
    
    wb.save(filename)


if __name__ == "__main__":
    csv_path = "/usr/openv/scripts/nbuchk/csv"
    src_data = get_csv_data(csv_path)
    date_day = datetime.date.today().strftime("%Y%m%d")
    if not src_data:
        print("Find csv files failed.")
        sys.exit()
    xlsx_filename = "/usr/openv/scripts/nbuchk/out/nbu_report_yz_" + date_day +".xlsx"
    generate_report(src_data,xlsx_filename)
